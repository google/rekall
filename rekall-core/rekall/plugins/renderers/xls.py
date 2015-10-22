# Rekall Memory Forensics
# Copyright 2014 Google Inc. All Rights Reserved.
#
# Authors:
# Michael Cohen <scudette@google.com>
#
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or (at
# your option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
# General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA 02111-1307 USA
#

"""This file implements an xls renderer based on the openpyxl project.

We produce xls (Excel spreadsheet files) with the output from Rekall plugins.
"""
import time
import openpyxl

from openpyxl import styles
from openpyxl.styles import colors
from openpyxl.styles import fills

from rekall import utils
from rekall.ui import renderer
from rekall.ui import text

# pylint: disable=unexpected-keyword-arg,no-value-for-parameter
# pylint: disable=redefined-outer-name

HEADER_STYLE = styles.Style(font=styles.Font(bold=True))
SECTION_STYLE = styles.Style(
    fill=styles.PatternFill(
        fill_type=fills.FILL_SOLID, start_color=styles.Color(colors.RED)))
FORMAT_STYLE = styles.Style(
    alignment=styles.Alignment(vertical="top", wrap_text=False))


class XLSObjectRenderer(renderer.ObjectRenderer):
    """By default the XLS renderer delegates to the text renderer."""
    renders_type = "object"
    renderers = ["XLSRenderer"]

    STYLE = None

    def _GetDelegateObjectRenderer(self, item):
        return self.ForTarget(item, "TextRenderer")(
            session=self.session, renderer=self.renderer.delegate_text_renderer)

    def RenderHeader(self, worksheet, column):
        cell = worksheet.cell(
            row=worksheet.current_row, column=worksheet.current_column)

        cell.value = column.name
        cell.style = HEADER_STYLE

        # Advance the pointer by 1 cell.
        worksheet.current_column += 1

    def RenderCell(self, value, worksheet, **options):
        # By default just render a single value into the current cell.
        cell = worksheet.cell(
            row=worksheet.current_row, column=worksheet.current_column)
        cell.value = self.GetData(value, **options)
        if self.STYLE:
            cell.style = self.STYLE

        # Advance the pointer by 1 cell.
        worksheet.current_column += 1

    def GetData(self, value, **options):
        if isinstance(value, (int, float, long)):
            return value

        return unicode(self._GetDelegateObjectRenderer(value).render_row(
            value, **options))


class XLSColumn(text.TextColumn):

    def __init__(self, type=None, table=None, renderer=None, session=None,
                 **options):
        super(XLSColumn, self).__init__(table=table, renderer=renderer,
                                        session=session, **options)

        if type:
            self.object_renderer = self.renderer.get_object_renderer(
                type=type, target_renderer="XLSRenderer", **options)


class XLSTable(text.TextTable):
    column_class = XLSColumn

    def render_header(self):
        current_ws = self.renderer.current_ws
        for column in self.columns:
            if column.object_renderer:
                object_renderer = column.object_renderer
            else:
                object_renderer = XLSObjectRenderer(
                    session=self.session, renderer=self.renderer)

            object_renderer.RenderHeader(self.renderer.current_ws, column)

        current_ws.current_row += 1
        current_ws.current_column = 1

    def render_row(self, row=None, highlight=None, **options):
        merged_opts = self.options.copy()
        merged_opts.update(options)

        # Get each column to write its own header.
        current_ws = self.renderer.current_ws
        for item in row:
            # Get the object renderer for the item.
            object_renderer = self.renderer.get_object_renderer(
                target=item, type=merged_opts.get("type"), **merged_opts)

            object_renderer.RenderCell(item, current_ws, **options)

        current_ws.current_row += 1
        current_ws.current_column = 1


class XLSRenderer(renderer.BaseRenderer):
    """A Renderer for xls files."""

    name = "xls"

    table_class = XLSTable
    tablesep = ""

    def __init__(self, output=None, **kwargs):
        super(XLSRenderer, self).__init__(**kwargs)

        # Make a single delegate text renderer for reuse. Most of the time we
        # will just replicate the output from the TextRenderer inside the
        # spreadsheet cell.
        self.delegate_text_renderer = text.TextRenderer(session=self.session)

        self.output = output or self.session.GetParameter("output")

        # If no output filename was give, just make a name based on the time
        # stamp.
        if self.output == None:
            self.output = "%s.xls" % time.ctime()

        try:
            self.wb = openpyxl.load_workbook(self.output)
            self.current_ws = self.wb.create_sheet()
        except IOError:
            self.wb = openpyxl.Workbook()
            self.current_ws = self.wb.active

    def start(self, plugin_name=None, kwargs=None):
        super(XLSRenderer, self).start(plugin_name=plugin_name, kwargs=kwargs)

        # Make a new worksheet for this run.
        if self.current_ws is None:
            self.current_ws = self.wb.create_sheet()

        ws = self.current_ws
        ws.title = plugin_name or ""
        ws.current_row = 1
        ws.current_column = 1

        return self

    def flush(self):
        super(XLSRenderer, self).flush()
        self.current_ws = None
        # Write the spreadsheet to a file.
        self.wb.save(self.output)

    def section(self, name=None, **_):
        ws = self.current_ws
        for i in range(10):
            cell = ws.cell(row=ws.current_row, column=i + 1)
            if i == 0:
                cell.value = name

            cell.style = SECTION_STYLE

        ws.current_row += 1
        ws.current_column = 1

    def format(self, formatstring, *data):
        worksheet = self.current_ws
        if "%" in formatstring:
            data = formatstring % data
        else:
            data = formatstring.format(*data)

        cell = worksheet.cell(
            row=worksheet.current_row, column=worksheet.current_column)
        cell.value = data
        cell.style = FORMAT_STYLE

        worksheet.current_column += 1
        if "\n" in data:
            worksheet.current_row += 1
            worksheet.current_column = 1

    def table_header(self, *args, **options):
        super(XLSRenderer, self).table_header(*args, **options)

        self.table.render_header()


# Following here are object specific renderers.


class XLSEProcessRenderer(XLSObjectRenderer):
    """Expands an EPROCESS into three columns (address, name and PID)."""
    renders_type = "_EPROCESS"

    def RenderHeader(self, worksheet, column):
        for heading in ["_EPROCESS", "Name", "PID"]:
            cell = worksheet.cell(
                row=worksheet.current_row, column=worksheet.current_column)
            cell.value = heading
            cell.style = HEADER_STYLE

            worksheet.current_column += 1

    def RenderCell(self, item, worksheet, **options):
        for value in ["%#x" % item.obj_offset, item.name, item.pid]:
            object_renderer = self.ForTarget(value, self.renderer)(
                session=self.session, renderer=self.renderer, **options)
            object_renderer.RenderCell(value, worksheet, **options)


class XLSStringRenderer(XLSObjectRenderer):
    renders_type = "String"

    def GetData(self, item, **_):
        return utils.SmartStr(item)


class XLSStructRenderer(XLSObjectRenderer):
    """Hex format struct's offsets."""
    renders_type = "Struct"

    def GetData(self, item, **_):
        return "%#x" % item.obj_offset


class XLSPointerRenderer(XLSObjectRenderer):
    """Renders the address of the pointer target as a hex string."""
    renders_type = "Pointer"

    def GetData(self, item, **_):
        result = item.v()
        if result == None:
            return "-"

        return "%#x" % result


class XLSNativeTypeRenderer(XLSObjectRenderer):
    """Renders native types as python objects."""
    renders_type = "NativeType"

    def GetData(self, item, **options):
        result = item.v()
        if result != None:
            return result


class XLS_UNICODE_STRING_Renderer(XLSNativeTypeRenderer):
    renders_type = "_UNICODE_STRING"


class XLSNoneObjectRenderer(XLSObjectRenderer):
    renders_type = "NoneObject"

    def GetData(self, item, **_):
        _ = item
        return "-"


class XLSDateTimeRenderer(XLSObjectRenderer):
    """Renders timestamps as python datetime objects."""
    renders_type = "UnixTimeStamp"
    STYLE = styles.Style(number_format='MM/DD/YYYY HH:MM:SS')

    def GetData(self, item, **options):
        if item.v() == 0:
            return None

        return item.as_datetime()
